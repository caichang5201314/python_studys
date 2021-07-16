'''
写入一个新的excel
'''
from openpyxl import Workbook
# 创建一个工作簿
workbook = Workbook()
# 创建一个表单
sheet = workbook.create_sheet('员工表',0)
# 写入一个数据
sheet.cell(row=1, column=1, value="python")
# 保存
workbook.save(r'd:\test.xlsx')