'''
读取指定单元格
'''
# 打开工作簿
from openpyxl import load_workbook

workbook = load_workbook(r'd:\test.xlsx')
# 获取表单
sheet = workbook['员工表']
# 读取指定的单元格数据
cell = sheet.cell(row=1, column=1).value
# 第二种读取单元格数据的方式
cell = sheet['A1'].value
print(cell)


# 读取A1-B4的单元格，共8个单元格
cell3 = sheet['A1':'B4']

# 读取A1-B4的单元格，共8个单元格
cell4 = sheet['A1:B4']

# 读取第2行的单元格
cell5 = sheet[2]

# 读取第1-2行的单元格
cell5 = sheet[1:2]
print(cell3[0][1].value)
print(cell4)
print(cell5)
