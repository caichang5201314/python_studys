# 1.openpyxl中列和行的起始标识都是1，不是从0开始
# 2.从excel中读取出来的数据只有两种类型，即数值类型和字符串类型
# 3.不要随便在表格中敲空格，会影响判断最大行数和最大列数
# 4.运行操作excel的代码时，要先关闭在操作系统中打开的相关excel表，否则可能会无法读取/写入数据
# 5.写入数据时要执行保存


import openpyxl

# 打开工作簿
wb = openpyxl.load_workbook(r'd:\test.xlsx')
# 获取表单
sh = wb['员工表']
# 读取指定的单元格数据
res1 = sh.cell(row=1, column=1).value
print(res1)

# 获取最大行数
print(sh.max_row)
# 获取最大列数
print(sh.max_column)

# 按列读取所有数据，每一列的单元格放入一个元组中
print(sh.columns)   # 直接打印，打印结果是一个可迭代对象，我们可以转换成列表来查看

# 按行读取所有数据，每一行的单元格放入一个元组中
rows = sh.rows
# print(list(rows))   # 转换成列表之后打印结果为具体的单元格，如下
# [(<Cell '表单1'.A1>, <Cell '表单1'.B1>, <Cell '表单1'.C1>, <Cell '表单1'.D1>, <Cell '表单1'.E1>, <Cell '表单1'.F1>, <Cell '表单1'.G1>),
#  (<Cell '表单1'.A2>, <Cell '表单1'.B2>, <Cell '表单1'.C2>, <Cell '表单1'.D2>, <Cell '表单1'.E2>, <Cell '表单1'.F2>, <Cell '表单1'.G2>),
#  (<Cell '表单1'.A3>, <Cell '表单1'.B3>, <Cell '表单1'.C3>, <Cell '表单1'.D3>, <Cell '表单1'.E3>, <Cell '表单1'.F3>, <Cell '表单1'.G3>)]


# 我们可以通过for循环以及value来查看单元格的值
for row in list(rows):  # 遍历每行数据
    case = []   # 用于存放一行数据
    for c in row:  # 把每行的每个单元格的值取出来，存放到case里
        case.append(c.value)
    print(case)

