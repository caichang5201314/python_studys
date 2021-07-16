from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class ExcelHelper():
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        """打开工作簿，选中表单"""
        self.wb = load_workbook(self.file_name)
        self.sh = self.wb[self.sheet_name]
        return self.sh

    def close(self):
        """关闭工作簿对象的方法"""
        self.wb.close()

    def read_by_cell_index(self, cell_index):
        """按行读取数据，最后返回一个存储字典的列表"""
        sheet = self.open()  # 打开工作簿
        cell = sheet[cell_index].value
        self.close()
        return cell

    def read_all(self,title_row=1):
        """按行读取数据，最后返回的是一个存储实例对象的列表"""
        all_data = []
        for item in list(self.sh.rows)[title_row:]:
            value_list = [l.value for l in item]
            all_data.append(value_list)
        return all_data

    def read_customer_area(self, start_index, end_index,title_row=1):
        """按行读取数据，最后返回的是一个存储实例对象的列表"""
        all_data = []
        for item in list(self.open()[start_index:end_index])[title_row:]:
            value_list = [l.value for l in item]
            all_data.append(value_list)
        return all_data

    def write_data(self, row, column, text):
        """写入数据的方法"""
        self.open()
        orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
        self.sh.cell(row=row, column=column, value=text)
        self.sh.cell(row=row,column=column).fill=orange_fill
        self.wb.save(self.file_name)
        self.close()


if __name__ == '__main__':
    # 直接运行本文件时执行，下面是一个应用实例
    # 需要读取excel时直接调用ReadExcel类
    test = ExcelHelper(r'd:\test.xlsx', '员工表')
    res = test.read_by_cell_index('C2')  # 最后返回的是一个存储实例对象的列表
    print(res)
    res = test.read_all()
    print(res)

    res = test.read_customer_area('A1','C5')
    print(res)

    #test.write_data(15,15,'caichang')
    print(test.file_name)
    test.write_data(10,10,'cai')